import json
import sqlite3
import openpyxl
from collections import OrderedDict

XLSX_PATH = r"K:\Code_folder\ohcard\mdb_scalesfixed.xlsx"
JSON_PATH = r"K:\Code_folder\ohcard\backend\db\mdb_scales.json"
DB_PATH   = r"K:\Code_folder\ohcard\backend\prisma\zhuojian.db"

wb = openpyxl.load_workbook(XLSX_PATH)

# ── 量表基本信息 ──────────────────────────────────────────────────────
ws1 = wb["量表信息"]
scales = OrderedDict()
for row in ws1.iter_rows(min_row=2, values_only=True):
    code = row[0]
    if not code:
        continue
    dims_str = row[10] or ""
    dims = [{"code": d.strip()} for d in dims_str.split("、") if d.strip()]
    method = row[11] if len(row) > 11 else None
    scoring_rule = {}
    if method:
        scoring_rule["method"] = method
    if dims:
        scoring_rule["dimensions"] = dims
    scales[code] = {
        "code": code,
        "name": row[1],
        "category": row[2],
        "description": row[3],
        "introduction": row[4],
        "instruction": row[5],
        "totalQuestions": row[6],
        "estimatedMinutes": row[7],
        "isPaid": row[8] == "是",
        "isActive": row[9] == "是",
        "scoringRule": scoring_rule,
        "questions": [],
        "interpretations": [],
    }

# ── 题目明细 ──────────────────────────────────────────────────────────
ws2 = wb["题目明细"]
for row in ws2.iter_rows(min_row=2, values_only=True):
    code = row[0]
    if not code or code not in scales:
        continue
    options = []
    for i, opt in enumerate((row[5] or "").split("  ")):
        opt = opt.strip()
        if "->" in opt:
            label, nxt = opt.split("->", 1)
            options.append({"value": i + 1, "label": label.strip(), "next": nxt.strip()})
        elif ":" in opt:
            v, label = opt.split(":", 1)
            try:
                options.append({"value": float(v), "label": label})
            except ValueError:
                options.append({"value": v, "label": label})
    scales[code]["questions"].append({
        "orderNum": row[2],
        "content": row[3],
        "dimension": row[4],
        "options": options,
    })

# ── 评价规则 ──────────────────────────────────────────────────────────
ws3 = wb["评价规则"]
interp_headers = [c.value for c in ws3[1]]
def col(name, fallback):
    try:
        return interp_headers.index(name)
    except ValueError:
        return fallback

level_col  = col('level', 13)

def to_float(v):
    try:
        return float(v) if v is not None else None
    except (TypeError, ValueError):
        return None

for row in ws3.iter_rows(min_row=2, values_only=True):
    code = row[0]
    if not code or code not in scales:
        continue
    scales[code]["interpretations"].append({
        "dimension": row[2],
        "calculationMethod": row[3],
        "standardScore": to_float(row[4]),
        "variance": to_float(row[5]),
        "minScore": row[6],
        "maxScore": row[7],
        "ageMin": row[8],
        "ageMax": row[9],
        "gender": row[10],
        "isNormal": row[11],
        "detail": row[12],
        "level": row[level_col] if len(row) > level_col else None,
    })

result = list(scales.values())

# 对 dimension_sum 量表，若 dim 无 questions，从题目 dimension 字段自动构建
for s in result:
    rule = s['scoringRule']
    if rule.get('method') == 'dimension_sum' and rule.get('dimensions'):
        if not any('questions' in d for d in rule['dimensions']):
            dim_map = {}
            for q in s['questions']:
                if q.get('dimension'):
                    dim_map.setdefault(q['dimension'], []).append(q['orderNum'])
            for d in rule['dimensions']:
                d['questions'] = dim_map.get(d['code'], [])

# ── 写 JSON ───────────────────────────────────────────────────────────
with open(JSON_PATH, "w", encoding="utf-8") as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

# ── 写 DB (upsert) ───────────────────────────────────────────────────
db = sqlite3.connect(DB_PATH)
cur = db.cursor()

inserted = updated = 0
for s in result:
    cur.execute("SELECT id FROM AssessmentScale WHERE code=?", (s["code"],))
    row = cur.fetchone()
    rule_json = json.dumps(s["scoringRule"], ensure_ascii=False)

    if row:
        scale_id = row[0]
        cur.execute("""
            UPDATE AssessmentScale SET
              name=?, category=?, description=?, introduction=?, instruction=?,
              isPaid=?, isActive=?, totalQuestions=?, estimatedMinutes=?, scoringRule=?
            WHERE id=?
        """, (
            s["name"], s["category"], s["description"], s.get("introduction"), s.get("instruction"),
            1 if s["isPaid"] else 0, 1 if s["isActive"] else 0,
            s["totalQuestions"], s["estimatedMinutes"], rule_json, scale_id
        ))
        cur.execute("DELETE FROM AssessmentQuestion WHERE scaleId=?", (scale_id,))
        cur.execute("DELETE FROM AssessmentInterpretation WHERE scaleId=?", (scale_id,))
        updated += 1
    else:
        cur.execute("""
            INSERT INTO AssessmentScale
              (code, name, category, description, introduction, instruction,
               isPaid, isActive, totalQuestions, estimatedMinutes, scoringRule, scenarios)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            s["code"], s["name"], s["category"], s["description"],
            s.get("introduction"), s.get("instruction"),
            1 if s["isPaid"] else 0, 1 if s["isActive"] else 0,
            s["totalQuestions"], s["estimatedMinutes"], rule_json, "[]"
        ))
        scale_id = cur.lastrowid
        inserted += 1

    for q in s["questions"]:
        cur.execute("""
            INSERT INTO AssessmentQuestion (scaleId, orderNum, content, options, dimension)
            VALUES (?,?,?,?,?)
        """, (scale_id, q["orderNum"], q["content"],
              json.dumps(q["options"], ensure_ascii=False), q.get("dimension")))

    for i in s["interpretations"]:
        cur.execute("""
            INSERT INTO AssessmentInterpretation
              (scaleId, dimension, calculationMethod, standardScore, variance,
               minScore, maxScore, ageMin, ageMax, gender, isNormal, level, detail)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            scale_id,
            i.get("dimension"), i.get("calculationMethod"),
            i.get("standardScore"), i.get("variance"),
            i["minScore"], i["maxScore"],
            i.get("ageMin"), i.get("ageMax"),
            i.get("gender"), i.get("isNormal"),
            i.get("level"), i["detail"]
        ))

db.commit()
db.close()
print(f"完成：新增 {inserted} 个，更新 {updated} 个量表")
