import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

with open(r"K:\Code_folder\ohcard\backend\db\mdb_scales.json", encoding="utf-8") as f:
    scales = json.load(f)

wb = openpyxl.Workbook()

# Sheet 1: 量表基本信息
ws1 = wb.active
ws1.title = "量表信息"
headers1 = ["编码", "名称", "分类", "描述", "简介", "引导词", "题目数", "预计分钟", "是否付费", "是否启用", "计算方式", "维度列表"]
ws1.append(headers1)
for cell in ws1[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="D9E1F2")

for s in scales:
    scoring_method = s.get("scoringRule", {}).get("method", "")
    dims = "、".join(d["code"] for d in s.get("scoringRule", {}).get("dimensions", []))
    ws1.append([
        s.get("code"), s.get("name"), s.get("category"), s.get("description"),
        s.get("introduction"), s.get("instruction"),
        s.get("totalQuestions"), s.get("estimatedMinutes"),
        "是" if s.get("isPaid") else "否",
        "是" if s.get("isActive") else "否",
        scoring_method, dims,
    ])

# Sheet 2: 题目明细
ws2 = wb.create_sheet("题目明细")
headers2 = ["量表编码", "量表名称", "题号", "题目内容", "维度", "选项（值:标签）"]
ws2.append(headers2)
for cell in ws2[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="D9E1F2")

for s in scales:
    for q in s.get("questions", []):
        options_str = "  ".join(f"{o['value']}:{o['label']}" for o in q.get("options", []))
        ws2.append([
            s.get("code"), s.get("name"),
            q.get("orderNum"), q.get("content"),
            q.get("dimension", ""), options_str,
        ])

# Sheet 3: 评价规则
ws3 = wb.create_sheet("评价规则")
headers3 = ["量表编码", "量表名称", "维度", "最低分", "最高分", "年龄下限", "年龄上限", "性别", "评价内容"]
ws3.append(headers3)
for cell in ws3[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="D9E1F2")

for s in scales:
    for interp in s.get("interpretations", []):
        ws3.append([
            s.get("code"), s.get("name"),
            interp.get("dimension"), interp.get("minScore"), interp.get("maxScore"),
            interp.get("ageMin"), interp.get("ageMax"),
            interp.get("gender"), interp.get("detail"),
        ])

# 自动列宽
for ws in [ws1, ws2, ws3]:
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

out = r"K:\Code_folder\ohcard\mdb_scales.xlsx"
wb.save(out)
print(f"已导出：{out}")
