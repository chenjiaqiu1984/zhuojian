import sqlite3
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

DB_PATH = "backend/prisma/zhuojian.db"
OUT_PATH = "assessment_export.xlsx"

conn = sqlite3.connect(DB_PATH)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

wb = openpyxl.Workbook()
wb.remove(wb.active)

HEADER_FILL = PatternFill("solid", fgColor="4A8A7A")
INTERP_FILL = PatternFill("solid", fgColor="3A6E80")
HEADER_FONT = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")

scales = cur.execute("SELECT * FROM AssessmentScale WHERE isActive=1 ORDER BY id").fetchall()

for scale in scales:
    sid = scale["id"]
    name = scale["name"][:31]  # sheet name max 31 chars
    ws = wb.create_sheet(title=name)

    # --- scoring method ---
    scoring_rule = json.loads(scale["scoringRule"] or "{}") if scale["scoringRule"] else {}
    method = scoring_rule.get("method", "")
    ws.append([f"计算方式: {method}"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, color="E07050")
    ws.append([])

    # --- questions ---
    ws.append(["题目", "选项", "维度/得分规则"])
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP

    questions = cur.execute(
        "SELECT * FROM AssessmentQuestion WHERE scaleId=? ORDER BY orderNum", (sid,)
    ).fetchall()

    for q in questions:
        try:
            opts = json.loads(q["options"])
            if isinstance(opts, list):
                opts_str = "\n".join(
                    f"{o.get('label','')}: {o.get('text', o.get('content',''))}  [{o.get('score','')}分]"
                    if isinstance(o, dict) else str(o)
                    for o in opts
                )
            else:
                opts_str = str(opts)
        except Exception:
            opts_str = q["options"]

        dim = q["dimension"] or ""
        row = ws.max_row + 1
        ws.cell(row, 1, f"Q{q['orderNum']}. {q['content']}")
        ws.cell(row, 2, opts_str)
        ws.cell(row, 3, dim)
        for c in [ws.cell(row, i) for i in range(1, 4)]:
            c.alignment = WRAP

    # --- interpretations ---
    interps = cur.execute(
        "SELECT * FROM AssessmentInterpretation WHERE scaleId=? ORDER BY dimension, minScore", (sid,)
    ).fetchall()

    if interps:
        ws.append([])
        header_row = ws.max_row + 1
        ws.append(["维度", "分数范围", "等级", "评判说明"])
        for cell in ws[header_row]:
            cell.fill = INTERP_FILL
            cell.font = HEADER_FONT
            cell.alignment = WRAP
        for i in interps:
            row = ws.max_row + 1
            ws.cell(row, 1, i["dimension"] or "总分")
            ws.cell(row, 2, f"{i['minScore']} ~ {i['maxScore']}")
            ws.cell(row, 3, i["level"] or "")
            ws.cell(row, 4, i["detail"])
            for c in [ws.cell(row, j) for j in range(1, 5)]:
                c.alignment = WRAP

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 50

conn.close()
wb.save(OUT_PATH)
print(f"导出完成: {OUT_PATH}，共 {len(wb.sheetnames)} 个量表")
