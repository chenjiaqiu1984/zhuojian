import pyodbc
import openpyxl
from openpyxl.styles import Font, PatternFill

DB_PATH = r"K:\Code_folder\ohcard\GUESTINFODB.mdb"
OUT_PATH = r"K:\Code_folder\ohcard\GUESTINFODB.xlsx"

conn = pyodbc.connect(
    rf"Driver={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={DB_PATH};"
)
cur = conn.cursor()

wb = openpyxl.Workbook()
wb.remove(wb.active)

FILL = PatternFill("solid", fgColor="4A8A7A")
FONT = Font(bold=True, color="FFFFFF")

def add_sheet(name, headers, rows):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = FILL
        cell.font = FONT
    for row in rows:
        ws.append([str(v) if v is not None else "" for v in row])
    for col in ws.columns:
        width = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 40)
    return ws

# Sheet 1: 受测者档案
rows = cur.execute("SELECT id, mzh, name, xb, csdate, hwno FROM usebb ORDER BY id").fetchall()
add_sheet("受测者档案", ["ID", "门诊号", "姓名", "性别", "出生日期", "病历号"], rows)

# Sheet 2: 就诊记录
rows = cur.execute("""
    SELECT b.mzh, b.name, k.hwno, k.hwdate, k.hcm, k.wkg, k.cx, k.tel, k.addr
    FROM usekb k LEFT JOIN usebb b ON k.mzh = b.mzh
    ORDER BY k.id
""").fetchall()
add_sheet("就诊记录", ["门诊号", "姓名", "病历号", "就诊日期", "科室", "职业", "婚状", "电话", "地址"], rows)

# Sheet 3: 联系人
rows = cur.execute("""
    SELECT g.mzh, b.name, g.name, g.ral, g.tel
    FROM usexg g LEFT JOIN usebb b ON g.mzh = b.mzh
    ORDER BY g.id
""").fetchall()
add_sheet("联系人", ["门诊号", "受测者", "联系人姓名", "关系", "电话"], rows)

# Sheet 4: 测试结果
rows = cur.execute("""
    SELECT j.id, j.userid, j.username, j.timeing, j.hmna, j.tmc,
           j.sumcount, j.avgcount, j.type, j.exlev, j.exc
    FROM jg j ORDER BY j.id
""").fetchall()
add_sheet("测试结果", ["ID", "用户ID", "操作员", "测试时间", "量表名称", "用时(秒)",
                      "总分", "均分", "类型", "等级", "说明"], rows)

# Sheet 5: 分量表明细
rows = cur.execute("""
    SELECT j.username, j.timeing, j.hmna, m.hmna, m.sumcount, m.avgcount, m.type, m.explain
    FROM jgmx m LEFT JOIN jg j ON m.jgid = j.id
    ORDER BY m.id
""").fetchall()
add_sheet("分量表明细", ["操作员", "测试时间", "量表", "分量表", "得分", "均分", "类型", "解释"], rows)

# Sheet 6: 测试记录
rows = cur.execute("SELECT id, mzh, hwno, hmlr, hmjg FROM usexx ORDER BY id").fetchall()
add_sheet("测试记录", ["ID", "门诊号", "病历号", "量表录入", "量表结果"], rows)

conn.close()
wb.save(OUT_PATH)
print(f"导出完成: {OUT_PATH}，共 {len(wb.sheetnames)} 个工作表")
