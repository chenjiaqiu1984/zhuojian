import pyodbc
import openpyxl
from openpyxl.styles import Font, PatternFill

DB_PATH = r"K:\Code_folder\ohcard\SYSINFODB.mdb"
OUT_PATH = r"K:\Code_folder\ohcard\SYSINFODB.xlsx"

conn = pyodbc.connect(
    rf"Driver={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={DB_PATH};PWD=empgao;"
)
cur = conn.cursor()

wb = openpyxl.Workbook()
wb.remove(wb.active)

FILL = PatternFill("solid", fgColor="3A6E80")
FONT = Font(bold=True, color="FFFFFF")

def add_sheet(name, headers, rows):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = FILL
        cell.font = FONT
    for row in rows:
        ws.append([str(v) if v is not None else "" for v in row])
    for col in ws.columns:
        width = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 40)

# 量表名称目录树
rows = cur.execute("SELECT id, name, parentid, chc, max, min, xb, hmjg FROM treename ORDER BY id").fetchall()
add_sheet("量表目录", ["ID", "名称", "父ID", "选项数", "最高分", "最低分", "性别", "结果规则"], rows)

# 评分区间及解释
rows = cur.execute("SELECT no, hmid, name, agemin, agemax, xb, min, max, explain FROM account ORDER BY hmid, min").fetchall()
add_sheet("评分区间", ["编号", "量表ID", "量表名", "年龄下限", "年龄上限", "性别", "最低分", "最高分", "解释"], rows)

# 常模
rows = cur.execute("SELECT Age, Gender, Average, SD FROM AverageSD ORDER BY Age, Gender").fetchall()
add_sheet("常模", ["年龄", "性别", "均值", "标准差"], rows)

# 分量表题目分组
rows = cur.execute("SELECT id, hmid, agemin, agemax, xb, ez, topid FROM hmez ORDER BY hmid, id").fetchall()
add_sheet("分量表分组", ["ID", "量表ID", "年龄下限", "年龄上限", "性别", "分量表名", "题目ID"], rows)

# 题目选项及分值（toplr）
rows = cur.execute("SELECT no, hmid, topid, hmna, toplr, chc, a,b,c,d,e,f,g,h,i,j FROM toplr ORDER BY hmid, topid").fetchall()
add_sheet("题目toplr", ["编号","量表ID","题目ID","量表名","题干","选项数","A","B","C","D","E","F","G","H","I","J"], rows)

# TT 表
rows = cur.execute("SELECT no, hmid, topid, hmna, toplr, chc, a,b,c,d,e,f,g,h,i,j FROM TT ORDER BY hmid, topid").fetchall()
add_sheet("题目TT", ["编号","量表ID","题目ID","量表名","题干","选项数","A","B","C","D","E","F","G","H","I","J"], rows)

# MMPI T分转换表
rows = cur.execute("SELECT ID, User_Sex, LStyle, Factor_Number, Original_Score, Converted_Score FROM Convert_MMPI ORDER BY Factor_Number, Original_Score").fetchall()
add_sheet("MMPI转换表", ["ID","性别","LStyle","因子编号","原始分","T分"], rows)

conn.close()
wb.save(OUT_PATH)
print(f"导出完成: {OUT_PATH}，共 {len(wb.sheetnames)} 个工作表")
